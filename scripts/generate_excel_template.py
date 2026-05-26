from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXPECTED_COLUMNS = 122
ROOT = Path(__file__).resolve().parents[1]
INDEX_HTML = ROOT / "index.html"
DICCIONARIO_MD = ROOT / "docs" / "DICCIONARIO_VARIABLES.md"
OUTPUT_XLSX = ROOT / "templates" / "BD_VISITAS_HS_template.xlsx"


def extract_master_columns(index_html: Path) -> List[str]:
    content = index_html.read_text(encoding="utf-8")
    pattern = r"const\s+MASTER_COLUMNS\s*=\s*\[(?P<body>[\s\S]*?)\];"
    match = re.search(pattern, content)
    if not match:
        raise ValueError("No se encontró el bloque MASTER_COLUMNS en index.html.")

    body = match.group("body")
    columns = re.findall(r"'([^']+)'", body)
    if not columns:
        raise ValueError("MASTER_COLUMNS existe, pero no contiene columnas parseables.")

    duplicates = sorted({c for c in columns if columns.count(c) > 1})
    if duplicates:
        raise ValueError(f"Se detectaron columnas duplicadas en MASTER_COLUMNS: {duplicates}")

    if len(columns) != EXPECTED_COLUMNS:
        raise ValueError(
            f"MASTER_COLUMNS tiene {len(columns)} columnas, pero se esperaban {EXPECTED_COLUMNS}."
        )

    return columns


def parse_dictionary_table(md_path: Path) -> Tuple[List[str], List[List[str]], List[str]]:
    lines = md_path.read_text(encoding="utf-8").splitlines()
    in_table = False
    header: List[str] = []
    rows: List[List[str]] = []
    warnings: List[str] = []

    for line in lines:
        if not in_table:
            if line.strip().startswith("| variable |"):
                in_table = True
                header = [cell.strip() for cell in line.strip().strip("|").split("|")]
            continue

        stripped = line.strip()
        if stripped.startswith("|---"):
            continue
        if not stripped.startswith("|"):
            if rows:
                break
            continue

        parts = [cell.strip() for cell in stripped.strip("|").split("|")]
        if len(parts) != len(header):
            warnings.append(f"Línea de tabla no parseada (columnas inesperadas): {line}")
            continue
        rows.append(parts)

    if not header or not rows:
        raise ValueError("No se pudo parsear la tabla principal del diccionario en Markdown.")

    return header, rows, warnings


def extract_proms_section(md_path: Path) -> List[str]:
    lines = md_path.read_text(encoding="utf-8").splitlines()
    start_idx = None
    for i, line in enumerate(lines):
        if line.strip() == "## PROMs: fuente, licencia y versionado interno":
            start_idx = i
            break

    if start_idx is None:
        return ["No se encontró sección PROMs en docs/DICCIONARIO_VARIABLES.md."]

    section: List[str] = []
    for line in lines[start_idx:]:
        if line.startswith("## ") and line.strip() != "## PROMs: fuente, licencia y versionado interno":
            break
        section.append(line)

    cleaned = [s.rstrip() for s in section if s.strip() != ""]
    return cleaned


def apply_header_style(ws, columns: List[str]) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col_idx, value in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(16, min(40, len(value) + 4))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def write_dictionary_sheet(ws, header: List[str], rows: List[List[str]], proms_lines: List[str]) -> None:
    for col_idx, value in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 28

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    start_proms_row = len(rows) + 4
    ws.cell(row=start_proms_row, column=1, value="SECCIÓN PROMs (fuente/licencia/versionado)").font = Font(bold=True)
    ws.merge_cells(start_row=start_proms_row, start_column=1, end_row=start_proms_row, end_column=len(header))

    current = start_proms_row + 1
    for line in proms_lines:
        ws.cell(row=current, column=1, value=line)
        ws.merge_cells(start_row=current, start_column=1, end_row=current, end_column=len(header))
        ws.cell(row=current, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        current += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"


def write_instructions_sheet(ws) -> None:
    lines = [
        "Plantilla BD_VISITAS_HS — Consulta Enfermería HS HUVR",
        "",
        "Esta plantilla recoge una fila por visita.",
        "Todas las visitas se pegan en la hoja BD_VISITAS_HS.",
        "No crear hojas separadas para PV, SG o CX.",
        "La columna tipo_visita identifica el tipo de registro:",
        "PV = Primera visita",
        "SG = Seguimiento",
        "CX = Cura postquirúrgica",
        "No modificar nombres de columnas.",
        "No reordenar columnas.",
        "No borrar columnas aunque estén vacías.",
        "Los campos que no aplican pueden quedar vacíos.",
        "codigo_hs está reservado para fase posterior.",
        "Esta plantilla es una herramienta de registro/análisis del piloto, no sustituye la historia clínica oficial.",
        "Guardar siempre el archivo en ubicación autorizada por el hospital/equipo.",
        "No introducir datos identificables en archivos compartidos fuera del entorno autorizado.",
    ]

    ws.column_dimensions["A"].width = 150
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if i == 1:
            cell.font = Font(bold=True, size=14)


def build_workbook(columns: List[str], dictionary_header: List[str], dictionary_rows: List[List[str]], proms_lines: List[str]) -> Workbook:
    wb = Workbook()

    ws_data = wb.active
    ws_data.title = "BD_VISITAS_HS"
    apply_header_style(ws_data, columns)

    ws_dict = wb.create_sheet("DICCIONARIO_VARIABLES")
    write_dictionary_sheet(ws_dict, dictionary_header, dictionary_rows, proms_lines)

    ws_inst = wb.create_sheet("INSTRUCCIONES")
    write_instructions_sheet(ws_inst)

    return wb


def main() -> int:
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    try:
        columns = extract_master_columns(INDEX_HTML)
        dictionary_header, dictionary_rows, warnings = parse_dictionary_table(DICCIONARIO_MD)
        proms_lines = extract_proms_section(DICCIONARIO_MD)
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1

    for warning in warnings:
        print(f"ADVERTENCIA: {warning}")

    wb = build_workbook(columns, dictionary_header, dictionary_rows, proms_lines)
    wb.save(OUTPUT_XLSX)

    print(f"OK: Plantilla generada en {OUTPUT_XLSX}")
    print(f"OK: Columnas detectadas: {len(columns)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
