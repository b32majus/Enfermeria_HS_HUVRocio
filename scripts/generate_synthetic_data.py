from __future__ import annotations

import re
import sys
from datetime import date
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook

from generate_excel_template import (
    DICCIONARIO_MD,
    EXPECTED_COLUMNS,
    INDEX_HTML,
    apply_header_style,
    extract_master_columns,
    extract_proms_section,
    parse_dictionary_table,
    write_dictionary_sheet,
    write_instructions_sheet,
)


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_XLSX = ROOT / "templates" / "BD_VISITAS_HS_sintetico.xlsx"
TOOL_VERSION = "v1.3-longitudinal-export"


def ihs4_score(nod: int, abscesos: int, fistulas: int) -> int:
    return nod + (2 * abscesos) + (4 * fistulas)


def ihs4_severity(score: int) -> str:
    if score <= 3:
        return "Leve"
    if score <= 10:
        return "Moderado"
    return "Grave"


def days_to_text(days: int) -> str:
    if days < 90:
        return f"{days} días"
    months = round(days / 30.44)
    return f"{months} meses aprox."


def iso(d: date) -> str:
    return d.isoformat()


def new_row(columns: List[str], overrides: Dict[str, str]) -> Dict[str, str]:
    row = {c: "" for c in columns}
    row.update({k: str(v) for k, v in overrides.items() if k in row and v is not None})
    return row


def set_ihs_fields(row: Dict[str, str], nod: int, abscesos: int, fistulas: int, zonas: str) -> None:
    score = ihs4_score(nod, abscesos, fistulas)
    row["nodulos_total"] = str(nod)
    row["abscesos_total"] = str(abscesos)
    row["fistulas_total"] = str(fistulas)
    row["ihs4_actual"] = str(score)
    row["gravedad_ihs4"] = ihs4_severity(score)
    zone_items = [z.strip() for z in zonas.split(",") if z.strip()]
    row["zonas_activas_n"] = str(len(zone_items))
    row["zonas_activas_listado"] = ", ".join(zone_items)


def base_fields(tipo: str, fecha_visita: date, nuhsa: str, codigo_hs: str) -> Dict[str, str]:
    return {
        "fecha_exportacion": f"{iso(fecha_visita)}T08:30:00",
        "version_herramienta": TOOL_VERSION,
        "tipo_visita": tipo,
        "fecha_visita": iso(fecha_visita),
        "nuhsa": nuhsa,
        "codigo_hs": codigo_hs,
    }


def build_rows(columns: List[str]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []

    # Paciente 1: PV + SG (mejoría clara)
    p1_pv_date = date(2026, 1, 10)
    p1_sg_date = date(2026, 2, 24)  # 45 días después
    p1_pv = new_row(columns, {
        **base_fields("PV", p1_pv_date, "TEST0001", "HS0001"),
        "edad": "31", "sexo": "Mujer", "situacion_laboral": "Activa",
        "anio_inicio_sintomas": "2019", "anio_diagnostico": "2021", "retraso_diagnostico": "2 años",
        "profesional_diagnostico": "Dermatología", "antecedentes_familiares_hs": "No",
        "brotes_ultimo_anio": "7", "n_medicos_previos": "3", "fiebre_brotes": "No",
        "urgencias_hs_ultimo_anio": "2", "cirugias_previas_hs": "0", "biologico_previo": "No",
        "tratamientos_previos_relevantes": "Doxiciclina intermitente",
        "gine_obst": "G0P0", "deseos_genesicos": "Sí",
        "tabaco_estado": "No", "cigarrillos_dia": "0", "anios_fumando": "0", "intentos_cesacion": "No",
        "alcohol_ube_semana": "1", "ejercicio": "Moderada (3-5h/sem)", "sueno": "Regular (despertares)",
        "peso_kg": "74", "talla_m": "1.66", "imc": "26.9", "perimetro_abdominal_cm": "89",
        "hta": "No", "dislipemia": "No", "diabetes": "No", "obesidad": "No",
        "acne_grave_conglobata": "No", "sinus_pilonidal": "No", "eii_digestivo": "No",
        "espa_osteoarticular": "No", "hurley_enfermeria": "II",
        "eva_dolor": "8", "eva_prurito": "6", "eva_supuracion": "7",
        "dlqi_total": "18", "hads_ansiedad_total": "10", "hads_depresion_total": "7",
        "hsqol24_total": "52", "hsqol24_interpretacion": "Severo",
        "educacion_sanitaria_realizada": "Qué es la HS; Higiene y ropa; Educación terapéutica",
        "material_educativo_entregado": "Sí",
        "necesidades_valorar_dermatologia": "Valorar inicio de biológico",
        "proxima_cita_enfermeria": iso(p1_sg_date),
        "notas_enfermeria": "Dolor elevado, se refuerzan autocuidados.",
    })
    set_ihs_fields(p1_pv, nod=7, abscesos=3, fistulas=1, zonas="Axila derecha, Ingle izquierda, Glúteo izquierdo")
    rows.append(p1_pv)

    p1_sg = new_row(columns, {
        **base_fields("SG", p1_sg_date, "TEST0001", "HS0001"),
        "hurley_enfermeria": "II", "ihs4_previo": p1_pv["ihs4_actual"],
        "eva_dolor": "4", "eva_prurito": "3", "eva_supuracion": "3",
        "fecha_ultima_consulta": iso(p1_pv_date), "tiempo_desde_ultima_consulta": days_to_text(45),
        "estado_global_referido": "Mejor", "brotes_desde_ultima_visita": "1",
        "tratamiento_realizado_desde_ultima": "Inicio adalimumab y cuidados locales",
        "tratamiento_activo": "Adalimumab", "suspension_prematura": "No", "motivo_suspension": "",
        "adherencia_morisky_score": "4", "adherencia_morisky_interpretacion": "Alta",
        "adherencia_morisky_patron": "Completo", "efectos_adversos": "Ninguno",
        "detalle_efectos_adversos": "", "dlqi_total": "9", "hads_ansiedad_total": "6",
        "hads_depresion_total": "5", "hsqol24_total": "34", "hsqol24_interpretacion": "Moderado",
        "refuerzo_educativo_realizado": "Sí", "cura_hoy": "No", "tipo_herida": "", "aposito_cura": "",
        "notas_cura": "", "necesidades_valorar_dermatologia": "Mantener plan si continúa mejoría",
        "proxima_cita_enfermeria": iso(date(2026, 4, 15)),
        "notas_enfermeria": "Mejoría clínica y buena adherencia.",
    })
    set_ihs_fields(p1_sg, nod=2, abscesos=1, fistulas=0, zonas="Axila derecha, Ingle izquierda")
    rows.append(p1_sg)

    # Paciente 2: PV + SG + SG (fluctuante)
    p2_pv_date = date(2026, 1, 15)
    p2_sg1_date = date(2026, 3, 18)  # 62 días
    p2_sg2_date = date(2026, 6, 20)  # 94 días (~3 meses aprox)
    p2_pv = new_row(columns, {
        **base_fields("PV", p2_pv_date, "TEST0002", "HS0002"),
        "edad": "44", "sexo": "Hombre", "situacion_laboral": "Activa",
        "anio_inicio_sintomas": "2012", "anio_diagnostico": "2018", "retraso_diagnostico": "6 años",
        "profesional_diagnostico": "Dermatología", "antecedentes_familiares_hs": "Sí",
        "brotes_ultimo_anio": "9", "n_medicos_previos": "4", "fiebre_brotes": "No",
        "urgencias_hs_ultimo_anio": "1", "cirugias_previas_hs": "1", "biologico_previo": "Adalimumab",
        "tratamientos_previos_relevantes": "Clindamicina + Rifampicina",
        "gine_obst": "No aplica", "deseos_genesicos": "No aplica",
        "tabaco_estado": "Activo", "cigarrillos_dia": "10", "anios_fumando": "20", "intentos_cesacion": "1",
        "alcohol_ube_semana": "4", "ejercicio": "Ligera (paseos)", "sueno": "Mala (dolor nocturno)",
        "peso_kg": "98", "talla_m": "1.75", "imc": "32.0", "perimetro_abdominal_cm": "108",
        "hta": "Sí", "dislipemia": "Sí", "diabetes": "No", "obesidad": "Sí",
        "acne_grave_conglobata": "No", "sinus_pilonidal": "Sí", "eii_digestivo": "No",
        "espa_osteoarticular": "Sí", "hurley_enfermeria": "III",
        "eva_dolor": "7", "eva_prurito": "5", "eva_supuracion": "7",
        "dlqi_total": "16", "hads_ansiedad_total": "8", "hads_depresion_total": "6",
        "hsqol24_total": "46", "hsqol24_interpretacion": "Severo",
        "educacion_sanitaria_realizada": "Consejo antitabáquico; Nutrición / Peso; Educación terapéutica",
        "material_educativo_entregado": "Sí",
        "necesidades_valorar_dermatologia": "Valorar ajuste por persistencia de brotes",
        "proxima_cita_enfermeria": iso(p2_sg1_date),
        "notas_enfermeria": "Comorbilidad metabólica y tabaquismo activo.",
    })
    set_ihs_fields(p2_pv, nod=5, abscesos=2, fistulas=1, zonas="Axila derecha, Axila izquierda, Ingle derecha")
    rows.append(p2_pv)

    p2_sg1 = new_row(columns, {
        **base_fields("SG", p2_sg1_date, "TEST0002", "HS0002"),
        "hurley_enfermeria": "III", "ihs4_previo": p2_pv["ihs4_actual"],
        "eva_dolor": "6", "eva_prurito": "4", "eva_supuracion": "6",
        "fecha_ultima_consulta": iso(p2_pv_date), "tiempo_desde_ultima_consulta": days_to_text(62),
        "estado_global_referido": "Igual", "brotes_desde_ultima_visita": "2",
        "tratamiento_realizado_desde_ultima": "Continuación biológico + control local",
        "tratamiento_activo": "Adalimumab; Doxiciclina", "suspension_prematura": "No", "motivo_suspension": "",
        "adherencia_morisky_score": "3", "adherencia_morisky_interpretacion": "Intermedia",
        "adherencia_morisky_patron": "Completo", "efectos_adversos": "Fatiga",
        "detalle_efectos_adversos": "Fatiga leve al día siguiente de administración",
        "dlqi_total": "15", "hads_ansiedad_total": "8", "hads_depresion_total": "6",
        "hsqol24_total": "44", "hsqol24_interpretacion": "Severo",
        "refuerzo_educativo_realizado": "Sí", "cura_hoy": "No", "tipo_herida": "", "aposito_cura": "",
        "notas_cura": "", "necesidades_valorar_dermatologia": "Valorar escalado terapéutico si no mejora",
        "proxima_cita_enfermeria": iso(p2_sg2_date),
        "notas_enfermeria": "Evolución estable, adherencia parcial.",
    })
    set_ihs_fields(p2_sg1, nod=4, abscesos=2, fistulas=1, zonas="Axila derecha, Ingle derecha, Perianal")
    rows.append(p2_sg1)

    p2_sg2 = new_row(columns, {
        **base_fields("SG", p2_sg2_date, "TEST0002", "HS0002"),
        "hurley_enfermeria": "III", "ihs4_previo": p2_sg1["ihs4_actual"],
        "eva_dolor": "5", "eva_prurito": "4", "eva_supuracion": "5",
        "fecha_ultima_consulta": iso(p2_sg1_date), "tiempo_desde_ultima_consulta": days_to_text(94),
        "estado_global_referido": "Mejor", "brotes_desde_ultima_visita": "1",
        "tratamiento_realizado_desde_ultima": "Optimización pauta y refuerzo autocuidados",
        "tratamiento_activo": "Adalimumab", "suspension_prematura": "No", "motivo_suspension": "",
        "adherencia_morisky_score": "4", "adherencia_morisky_interpretacion": "Alta",
        "adherencia_morisky_patron": "Completo", "efectos_adversos": "Ninguno",
        "detalle_efectos_adversos": "", "dlqi_total": "11", "hads_ansiedad_total": "6",
        "hads_depresion_total": "5", "hsqol24_total": "36", "hsqol24_interpretacion": "Moderado",
        "refuerzo_educativo_realizado": "Sí", "cura_hoy": "Sí", "tipo_herida": "Fístula drenante",
        "aposito_cura": "Espuma absorbente", "notas_cura": "Exudado moderado, sin mal olor",
        "necesidades_valorar_dermatologia": "Mantener seguimiento estrecho",
        "proxima_cita_enfermeria": iso(date(2026, 8, 5)),
        "notas_enfermeria": "Mejoría parcial mantenida.",
    })
    set_ihs_fields(p2_sg2, nod=3, abscesos=1, fistulas=1, zonas="Axila derecha, Perianal")
    rows.append(p2_sg2)

    # Paciente 3: PV + CX
    p3_pv_date = date(2026, 2, 5)
    p3_cx_date = date(2026, 3, 1)
    p3_pv = new_row(columns, {
        **base_fields("PV", p3_pv_date, "TEST0003", "HS0003"),
        "edad": "29", "sexo": "Mujer", "situacion_laboral": "Activa",
        "anio_inicio_sintomas": "2020", "anio_diagnostico": "2022", "retraso_diagnostico": "2 años",
        "profesional_diagnostico": "Dermatología", "antecedentes_familiares_hs": "No",
        "brotes_ultimo_anio": "5", "n_medicos_previos": "2", "fiebre_brotes": "No",
        "urgencias_hs_ultimo_anio": "1", "cirugias_previas_hs": "0", "biologico_previo": "No",
        "tratamientos_previos_relevantes": "Resorcinol tópico",
        "gine_obst": "G1P1", "deseos_genesicos": "No",
        "tabaco_estado": "Ex", "cigarrillos_dia": "0", "anios_fumando": "8", "intentos_cesacion": "2-3",
        "alcohol_ube_semana": "0", "ejercicio": "Moderada (3-5h/sem)", "sueno": "Regular (despertares)",
        "peso_kg": "67", "talla_m": "1.64", "imc": "24.9", "perimetro_abdominal_cm": "82",
        "hta": "No", "dislipemia": "No", "diabetes": "No", "obesidad": "No",
        "acne_grave_conglobata": "No", "sinus_pilonidal": "No", "eii_digestivo": "No",
        "espa_osteoarticular": "No", "hurley_enfermeria": "II",
        "eva_dolor": "6", "eva_prurito": "5", "eva_supuracion": "6",
        "dlqi_total": "13", "hads_ansiedad_total": "7", "hads_depresion_total": "5",
        "hsqol24_total": "38", "hsqol24_interpretacion": "Moderado",
        "educacion_sanitaria_realizada": "Qué es la HS; Autocuidado heridas",
        "material_educativo_entregado": "Sí",
        "necesidades_valorar_dermatologia": "Valorar opción quirúrgica localizada",
        "proxima_cita_enfermeria": iso(p3_cx_date),
        "notas_enfermeria": "Lesión inguinal persistente, candidata a cirugía.",
    })
    set_ihs_fields(p3_pv, nod=4, abscesos=1, fistulas=1, zonas="Ingle derecha, Ingle izquierda")
    rows.append(p3_pv)

    p3_cx = new_row(columns, {
        **base_fields("CX", p3_cx_date, "TEST0003", "HS0003"),
        "n_cura": "1", "fecha_intervencion": iso(date(2026, 2, 27)), "cirujano": "Equipo QX HS",
        "tipo_intervencion": "Deroofing", "tipo_cierre": "Segunda intención", "localizacion_qx": "Inguinal derecha",
        "eva_pre_cura": "6", "eva_post_cura": "3", "analgesia_pre_cura": "Paracetamol",
        "anestesia_local": "Lidocaína 2% infiltrada",
        "hstime": "T(Tejido): Granulación sana (rojo) | I(Infección): Sin signos infección | M(Humedad): Exudado leve | E(Bordes): Avance epitelial activo",
        "herida_dimensiones": "3.2x1.5x0.8", "herida_color_lecho": "Rojo (granulación)", "herida_olor": "Ausente",
        "piel_perilesional": "Íntegra", "sangrado": "Leve", "solucion_limpieza": "Suero fisiológico",
        "desbridamiento": "No precisa", "lavado_h2o2": "No", "aposito_primario": "Hidrofibra (Aquacel)",
        "aposito_secundario": "Espuma absorbente", "fijacion": "Malla tubular", "tpn": "No",
        "complicaciones": "Ninguna", "detalle_complicaciones": "", "foto_clinica": "Sí (consentida)",
        "proxima_cura": iso(date(2026, 3, 4)), "frecuencia_cura": "Cada 48h",
        "notas_enfermeria": "Evolución postquirúrgica favorable.",
    })
    rows.append(p3_cx)

    # Paciente 4: solo PV
    p4_pv = new_row(columns, {
        **base_fields("PV", date(2026, 4, 3), "TEST0004", "HS0004"),
        "edad": "36", "sexo": "Mujer", "situacion_laboral": "Activa",
        "anio_inicio_sintomas": "2024", "anio_diagnostico": "2025", "retraso_diagnostico": "1 año",
        "profesional_diagnostico": "Atención Primaria", "antecedentes_familiares_hs": "Desconocido",
        "brotes_ultimo_anio": "3", "n_medicos_previos": "1", "fiebre_brotes": "No",
        "urgencias_hs_ultimo_anio": "0", "cirugias_previas_hs": "0", "biologico_previo": "No",
        "tratamientos_previos_relevantes": "Ninguno", "gine_obst": "G0P0", "deseos_genesicos": "Sí",
        "tabaco_estado": "No", "cigarrillos_dia": "0", "anios_fumando": "0", "intentos_cesacion": "No",
        "alcohol_ube_semana": "1", "ejercicio": "Ligera (paseos)", "sueno": "Buena",
        "peso_kg": "62", "talla_m": "1.68", "imc": "22.0", "perimetro_abdominal_cm": "77",
        "hta": "No", "dislipemia": "No", "diabetes": "No", "obesidad": "No",
        "acne_grave_conglobata": "No", "sinus_pilonidal": "No", "eii_digestivo": "No",
        "espa_osteoarticular": "No", "hurley_enfermeria": "I",
        "eva_dolor": "3", "eva_prurito": "2", "eva_supuracion": "3",
        "dlqi_total": "5", "hads_ansiedad_total": "4", "hads_depresion_total": "3",
        "hsqol24_total": "22", "hsqol24_interpretacion": "Sin impacto",
        "educacion_sanitaria_realizada": "Qué es la HS; Higiene y ropa",
        "material_educativo_entregado": "Sí",
        "necesidades_valorar_dermatologia": "Sin necesidades urgentes",
        "proxima_cita_enfermeria": iso(date(2026, 6, 10)),
        "notas_enfermeria": "Caso inicial leve, se programa seguimiento.",
    })
    set_ihs_fields(p4_pv, nod=2, abscesos=0, fistulas=0, zonas="Axila izquierda")
    rows.append(p4_pv)

    # Paciente 5: SG sin PV en este Excel
    p5_sg = new_row(columns, {
        **base_fields("SG", date(2026, 5, 2), "TEST0005", "HS0005"),
        "hurley_enfermeria": "II", "ihs4_previo": "9",
        "eva_dolor": "5", "eva_prurito": "4", "eva_supuracion": "5",
        "fecha_ultima_consulta": "2025-11-20", "tiempo_desde_ultima_consulta": "5 meses aprox.",
        "estado_global_referido": "Igual", "brotes_desde_ultima_visita": "2",
        "tratamiento_realizado_desde_ultima": "Seguimiento externo previo al piloto",
        "tratamiento_activo": "Secukinumab", "suspension_prematura": "No", "motivo_suspension": "",
        "adherencia_morisky_score": "3", "adherencia_morisky_interpretacion": "Intermedia",
        "adherencia_morisky_patron": "Completo", "efectos_adversos": "Cefalea",
        "detalle_efectos_adversos": "Cefalea autolimitada",
        "dlqi_total": "12", "hads_ansiedad_total": "7", "hads_depresion_total": "6",
        "hsqol24_total": "39", "hsqol24_interpretacion": "Moderado",
        "refuerzo_educativo_realizado": "Sí", "cura_hoy": "No", "tipo_herida": "", "aposito_cura": "",
        "notas_cura": "", "necesidades_valorar_dermatologia": "Revisar respuesta terapéutica",
        "proxima_cita_enfermeria": iso(date(2026, 6, 20)),
        "notas_enfermeria": "Paciente ya conocido de etapa previa.",
    })
    set_ihs_fields(p5_sg, nod=4, abscesos=1, fistulas=1, zonas="Axila derecha, Ingle izquierda")
    rows.append(p5_sg)

    # Paciente 6: PV + SG (empeoramiento)
    p6_pv_date = date(2026, 1, 28)
    p6_sg_date = date(2026, 4, 1)  # 63 días
    p6_pv = new_row(columns, {
        **base_fields("PV", p6_pv_date, "TEST0006", "HS0006"),
        "edad": "40", "sexo": "Hombre", "situacion_laboral": "Desempleo",
        "anio_inicio_sintomas": "2016", "anio_diagnostico": "2019", "retraso_diagnostico": "3 años",
        "profesional_diagnostico": "Dermatología", "antecedentes_familiares_hs": "Sí",
        "brotes_ultimo_anio": "6", "n_medicos_previos": "2", "fiebre_brotes": "Sí",
        "urgencias_hs_ultimo_anio": "3", "cirugias_previas_hs": "1", "biologico_previo": "No",
        "tratamientos_previos_relevantes": "Doxiciclina; Clindamicina + Rifampicina",
        "gine_obst": "No aplica", "deseos_genesicos": "No aplica",
        "tabaco_estado": "Activo", "cigarrillos_dia": "15", "anios_fumando": "18", "intentos_cesacion": "No",
        "alcohol_ube_semana": "5", "ejercicio": "Sedentario", "sueno": "Insomnio",
        "peso_kg": "104", "talla_m": "1.78", "imc": "32.8", "perimetro_abdominal_cm": "112",
        "hta": "Sí", "dislipemia": "Sí", "diabetes": "Sí", "obesidad": "Sí",
        "acne_grave_conglobata": "Sí", "sinus_pilonidal": "Sí", "eii_digestivo": "Sí",
        "espa_osteoarticular": "Sí", "hurley_enfermeria": "III",
        "eva_dolor": "6", "eva_prurito": "5", "eva_supuracion": "6",
        "dlqi_total": "17", "hads_ansiedad_total": "9", "hads_depresion_total": "8",
        "hsqol24_total": "47", "hsqol24_interpretacion": "Severo",
        "educacion_sanitaria_realizada": "Consejo antitabáquico; Nutrición / Peso; Educación terapéutica",
        "material_educativo_entregado": "Sí",
        "necesidades_valorar_dermatologia": "Valorar intensificación terapéutica",
        "proxima_cita_enfermeria": iso(p6_sg_date),
        "notas_enfermeria": "Perfil inflamatorio y metabólico complejo.",
    })
    set_ihs_fields(p6_pv, nod=4, abscesos=2, fistulas=1, zonas="Axila derecha, Axila izquierda, Perianal")
    rows.append(p6_pv)

    p6_sg = new_row(columns, {
        **base_fields("SG", p6_sg_date, "TEST0006", "HS0006"),
        "hurley_enfermeria": "III", "ihs4_previo": p6_pv["ihs4_actual"],
        "eva_dolor": "8", "eva_prurito": "7", "eva_supuracion": "8",
        "fecha_ultima_consulta": iso(p6_pv_date), "tiempo_desde_ultima_consulta": days_to_text(63),
        "estado_global_referido": "Peor", "brotes_desde_ultima_visita": "4",
        "tratamiento_realizado_desde_ultima": "Antibiótico oral sin respuesta suficiente",
        "tratamiento_activo": "Doxiciclina", "suspension_prematura": "Sí",
        "motivo_suspension": "Ineficacia clínica",
        "adherencia_morisky_score": "2", "adherencia_morisky_interpretacion": "Intermedia",
        "adherencia_morisky_patron": "Completo", "efectos_adversos": "GI (náuseas, diarrea)",
        "detalle_efectos_adversos": "Náuseas y diarrea intermitente",
        "dlqi_total": "21", "hads_ansiedad_total": "11", "hads_depresion_total": "9",
        "hsqol24_total": "58", "hsqol24_interpretacion": "Severo",
        "refuerzo_educativo_realizado": "Sí", "cura_hoy": "Sí", "tipo_herida": "Absceso drenado",
        "aposito_cura": "Alginato + espuma", "notas_cura": "Exudado abundante, dolor elevado",
        "necesidades_valorar_dermatologia": "Priorizar revisión terapéutica",
        "proxima_cita_enfermeria": iso(date(2026, 5, 5)),
        "notas_enfermeria": "Empeoramiento clínico y aumento de brotes.",
    })
    set_ihs_fields(p6_sg, nod=6, abscesos=3, fistulas=2, zonas="Axila derecha, Axila izquierda, Ingle derecha, Perianal")
    rows.append(p6_sg)

    rows.sort(key=lambda r: (r["nuhsa"], r["fecha_visita"]))
    return rows


def validate_rows(columns: List[str], rows: List[Dict[str, str]]) -> None:
    if len(columns) != EXPECTED_COLUMNS:
        raise ValueError(f"Se esperaban {EXPECTED_COLUMNS} columnas y se encontraron {len(columns)}.")

    duplicates = sorted({c for c in columns if columns.count(c) > 1})
    if duplicates:
        raise ValueError(f"Columnas duplicadas detectadas: {duplicates}")

    for i, row in enumerate(rows, start=1):
        if set(row.keys()) != set(columns):
            raise ValueError(f"Fila {i} no coincide exactamente con el esquema de {EXPECTED_COLUMNS} columnas.")
        if len(row) != EXPECTED_COLUMNS:
            raise ValueError(f"Fila {i} tiene {len(row)} columnas, esperado {EXPECTED_COLUMNS}.")

        nuhsa = row["nuhsa"]
        codigo = row["codigo_hs"]
        if not nuhsa.startswith("TEST"):
            raise ValueError(f"Fila {i}: NUHSA no sintético ({nuhsa}).")
        if not re.fullmatch(r"HS\d{4}", codigo):
            raise ValueError(f"Fila {i}: codigo_hs inválido ({codigo}).")
        if row["tipo_visita"] not in {"PV", "SG", "CX"}:
            raise ValueError(f"Fila {i}: tipo_visita inválido ({row['tipo_visita']}).")

        if row["tipo_visita"] in {"PV", "SG"} and row["ihs4_actual"]:
            n = int(row["nodulos_total"])
            a = int(row["abscesos_total"])
            f = int(row["fistulas_total"])
            score = ihs4_score(n, a, f)
            if int(row["ihs4_actual"]) != score:
                raise ValueError(f"Fila {i}: IHS4 incoherente.")
            if row["gravedad_ihs4"] != ihs4_severity(score):
                raise ValueError(f"Fila {i}: gravedad IHS4 incoherente.")

            listed = [z.strip() for z in row["zonas_activas_listado"].split(",") if z.strip()]
            if int(row["zonas_activas_n"]) != len(listed):
                raise ValueError(f"Fila {i}: zonas_activas_n no coincide con listado.")

    # SG con visita previa: fecha_ultima_consulta e ihs4_previo coherentes
    visits_by_patient: Dict[str, List[Dict[str, str]]] = {}
    for row in rows:
        visits_by_patient.setdefault(row["nuhsa"], []).append(row)

    for patient_rows in visits_by_patient.values():
        patient_rows.sort(key=lambda r: r["fecha_visita"])
        for idx, row in enumerate(patient_rows):
            if row["tipo_visita"] != "SG":
                continue
            if idx == 0:
                continue
            prev = patient_rows[idx - 1]
            if row["fecha_ultima_consulta"] != prev["fecha_visita"]:
                raise ValueError(
                    f"SG de {row['nuhsa']} con visita previa: fecha_ultima_consulta no coincide con visita anterior."
                )
            if row["ihs4_previo"] != prev["ihs4_actual"]:
                raise ValueError(
                    f"SG de {row['nuhsa']} con visita previa: ihs4_previo no coincide con ihs4_actual previo."
                )


def build_workbook(columns: List[str], rows: List[Dict[str, str]]) -> Workbook:
    header, dict_rows, warnings = parse_dictionary_table(DICCIONARIO_MD)
    for warning in warnings:
        print(f"ADVERTENCIA: {warning}")
    proms_lines = extract_proms_section(DICCIONARIO_MD)

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "BD_VISITAS_HS"
    apply_header_style(ws_data, columns)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=row[col])

    ws_dict = wb.create_sheet("DICCIONARIO_VARIABLES")
    write_dictionary_sheet(ws_dict, header, dict_rows, proms_lines)

    ws_inst = wb.create_sheet("INSTRUCCIONES")
    write_instructions_sheet(ws_inst)

    return wb


def main() -> int:
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    try:
        columns = extract_master_columns(INDEX_HTML)
        rows = build_rows(columns)
        validate_rows(columns, rows)
        wb = build_workbook(columns, rows)
        wb.save(OUTPUT_XLSX)
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1

    print(f"OK: Archivo sintético generado en {OUTPUT_XLSX}")
    print(f"OK: Pacientes sintéticos: {len({r['nuhsa'] for r in rows})}")
    print(f"OK: Filas sintéticas: {len(rows)}")
    print(f"OK: Columnas: {len(columns)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
